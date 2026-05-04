from __future__ import annotations

from copy import deepcopy
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

ELIGIBILITY_HEADERS = [
    "Member Name",
    "Member ID",
    "Status",
    "Status Reason",
    "Program",
    "Payer Code",
    "Coverage Category",
    "Billing Note",
    "EB01 Codes",
    "Service Types",
    "Benefit Entities",
    "Contacts",
    "AAA Codes",
    "ST Control #",
    "Trace #",
]

ERROR_HEADERS = [
    "Member Name",
    "Member ID",
    "Error Type",
    "AAA Code",
    "Error Summary",
    "Recommended Action",
    "Follow-up Action",
    "ST Control #",
    "Trace #",
]

PARSER_ISSUE_HEADERS = [
    "Transaction #",
    "Transaction Control #",
    "Segment",
    "Location",
    "Message",
    "Severity",
]


def test_export_xlsx_omits_parser_issues_sheet_for_clean_payload(
    client: TestClient,
) -> None:
    response = client.post("/api/v1/export/xlsx", json=_export_payload())

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Eligibility Results"]
    assert "Parser Issues" not in workbook.sheetnames
    summary_sheet = workbook["Summary"]
    assert summary_sheet.column_dimensions["A"].width == 18
    assert summary_sheet.column_dimensions["B"].width == 16
    assert summary_sheet["A1"].font.bold is True
    assert summary_sheet["A1"].fill.fill_type == "solid"
    assert summary_sheet["A1"].fill.fgColor.rgb in _rgb("F9FAFB")


def test_export_xlsx_includes_parser_issues_sheet_when_present(
    client: TestClient,
) -> None:
    payload = _export_payload()
    payload["parserIssueCount"] = 1
    payload["parserIssues"] = [
        {
            "transaction_index": 1,
            "transaction_control_number": "0002",
            "segment_id": "NM1",
            "location": "segment_position:999",
            "message": "Invalid benefit entity.",
            "severity": "error",
        }
    ]

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Eligibility Results", "Parser Issues"]
    assert "Parser Issues" in workbook.sheetnames
    sheet = workbook["Parser Issues"]
    assert [cell.value for cell in sheet[1]] == PARSER_ISSUE_HEADERS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:F2"
    assert sheet["B2"].value == "0002"
    assert sheet["B2"].number_format == "@"
    assert sheet["E2"].value == "Invalid benefit entity."
    assert sheet["E2"].alignment.wrap_text is True


def test_export_xlsx_adds_errors_sheet_for_not_found_even_when_error_count_is_zero(
    client: TestClient,
) -> None:
    payload = _export_payload()
    payload["summary"] = {
        "total": 1,
        "active": 0,
        "inactive": 0,
        "error": 0,
        "not_found": 1,
        "unknown": 0,
    }
    result = payload["results"][0]
    assert isinstance(result, dict)
    result["overall_status"] = "not_found"
    result["status_reason"] = "Subscriber not found"
    result["aaa_errors"] = []

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Errors", "Eligibility Results"]
    errors_sheet = workbook["Errors"]
    assert [cell.value for cell in errors_sheet[1]] == ERROR_HEADERS
    assert errors_sheet.freeze_panes == "A2"
    assert errors_sheet.auto_filter.ref == "A1:I2"
    assert errors_sheet["E2"].alignment.wrap_text is True
    assert errors_sheet["F2"].alignment.wrap_text is True
    assert errors_sheet["A2"].fill.fill_type is None
    assert errors_sheet["A2"].value == "DOE, JOHN"
    assert errors_sheet["C2"].value == "STATUS"
    assert errors_sheet["E2"].value == "Subscriber not found"
    assert errors_sheet["F2"].value == "Confirm the member ID and demographics before resubmitting."


def test_export_xlsx_errors_sheet_uses_aaa_message_and_suggestion(
    client: TestClient,
) -> None:
    payload = _export_payload()
    result = payload["results"][0]
    assert isinstance(result, dict)
    result["overall_status"] = "error"
    result["status_reason"] = "Invalid member ID."
    result["aaa_errors"] = [
        {
            "code": "72",
            "message": "Invalid member ID.",
            "follow_up_action_code": "C",
            "suggestion": "Confirm the member ID and retry.",
        }
    ]

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    errors_sheet = workbook["Errors"]
    assert errors_sheet["C2"].value == "AAA"
    assert errors_sheet["D2"].value == "72"
    assert errors_sheet["E2"].value == "Invalid member ID."
    assert errors_sheet["F2"].value == "Confirm the member ID and retry."
    assert errors_sheet["G2"].value == "C"


def test_export_xlsx_splits_pipe_delimited_plan_columns(client: TestClient) -> None:
    payload = _export_payload()
    result = payload["results"][0]
    assert isinstance(result, dict)
    segments = result["eligibility_segments"]
    assert isinstance(segments, list)
    segment = segments[0]
    assert isinstance(segment, dict)
    segment["plan_coverage_description"] = "DC MEDICAID FFS | 853Q | BUY-IN"

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Eligibility Results"]
    assert [cell.value for cell in sheet[1]] == ELIGIBILITY_HEADERS
    assert sheet["E2"].value == "DC MEDICAID FFS"
    assert sheet["F2"].value == "853Q"
    assert sheet["G2"].value == "BUY-IN"
    assert sheet["H2"].value == "Coverage on file"


def test_export_xlsx_plain_plan_description_stays_in_program_name(
    client: TestClient,
) -> None:
    response = client.post("/api/v1/export/xlsx", json=_export_payload())

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Eligibility Results"]
    assert sheet["E2"].value == "MEDICAID"
    assert sheet["F2"].value is None
    assert sheet["G2"].value is None


def test_export_xlsx_fills_error_rows_only(client: TestClient) -> None:
    payload = _export_payload()
    active_result = payload["results"][0]
    assert isinstance(active_result, dict)
    error_result = deepcopy(active_result)
    error_result["member_name"] = "ERROR, MEMBER"
    error_result["member_id"] = "99999999"
    error_result["overall_status"] = "error"
    error_result["status_reason"] = "Invalid member ID."
    error_result["aaa_errors"] = [
        {
            "code": "72",
            "message": "Invalid member ID.",
            "suggestion": "Confirm the member ID and retry.",
        }
    ]
    payload["results"] = [active_result, error_result]

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Eligibility Results"]
    assert sheet["A1"].fill.fill_type == "solid"
    assert sheet["A1"].fill.fgColor.rgb in _rgb("E8EEF7")
    assert sheet["A2"].fill.fill_type is None
    assert sheet["A3"].fill.fill_type == "solid"
    assert sheet["A3"].fill.fgColor.rgb in _rgb("FEF2F2")


def test_export_validation_xlsx_emits_three_sheets_with_matching_rows(
    client: TestClient,
) -> None:
    response = client.post("/api/v1/export/validation/xlsx", json=_validation_payload())

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Per-Patient", "Issues"]
    assert workbook["Summary"]["B5"].value == 3
    assert workbook["Per-Patient"].max_row == 4
    assert workbook["Issues"].max_row == 2
    assert workbook["Per-Patient"]["F3"].value == "invalid"
    assert workbook["Issues"]["C2"].value == "DCM_INVALID_PAYER_NAME"


def test_export_xlsx_applies_eligibility_table_layout_and_text_formats(
    client: TestClient,
) -> None:
    payload = _export_payload()
    result = payload["results"][0]
    assert isinstance(result, dict)
    result["member_id"] = "000123456"
    result["trace_number"] = "00000000000000012345"
    segments = result["eligibility_segments"]
    assert isinstance(segments, list)
    segment = segments[0]
    assert isinstance(segment, dict)
    segment["plan_coverage_description"] = "DC MEDICAID FFS | 853Q | BUY-IN"

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    sheet = workbook["Eligibility Results"]
    assert [cell.value for cell in sheet[1]] == ELIGIBILITY_HEADERS
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:O2"
    assert sheet.column_dimensions["D"].width == 34
    assert sheet.column_dimensions["E"].width == 32
    assert sheet["C2"].value == "Active"
    assert sheet["D2"].alignment.wrap_text is True
    assert sheet["E2"].alignment.wrap_text is True
    assert sheet["H2"].alignment.wrap_text is True
    for cell_ref in ("B2", "F2", "I2", "J2", "M2", "N2", "O2"):
        assert sheet[cell_ref].number_format == "@"


def _rgb(hex_rgb: str) -> set[str]:
    return {hex_rgb, f"00{hex_rgb}", f"FF{hex_rgb}"}


def _export_payload() -> dict[str, object]:
    return {
        "filename": "eligibility.xlsx",
        "payer_name": "DC MEDICAID",
        "summary": {
            "total": 1,
            "active": 1,
            "inactive": 0,
            "error": 0,
            "not_found": 0,
            "unknown": 0,
        },
        "results": [
            {
                "member_name": "DOE, JOHN",
                "member_id": "12345678",
                "overall_status": "active",
                "status_reason": "Coverage on file",
                "st_control_number": "0001",
                "trace_number": "TRACE0001",
                "eligibility_segments": [
                    {
                        "eligibility_code": "1",
                        "service_type_code": "30",
                        "service_type_codes": ["30"],
                        "plan_coverage_description": "MEDICAID",
                    }
                ],
                "benefit_entities": [
                    {
                        "entity_identifier_code": "P5",
                        "name": "PLAN SPONSOR",
                        "contacts": ["SUPPORT (TE:8665550001)"],
                    }
                ],
                "aaa_errors": [],
            }
        ],
    }


def _validation_payload() -> dict[str, object]:
    return {
        "filename": "validation.xlsx",
        "is_valid": False,
        "error_count": 1,
        "warning_count": 0,
        "issues": [
            {
                "severity": "error",
                "level": "dc_medicaid",
                "code": "DCM_INVALID_PAYER_NAME",
                "message": "Invalid payer name.",
                "transaction_index": 1,
                "transaction_control_number": "0002",
            }
        ],
        "patients": [
            {
                "index": 0,
                "transaction_control_number": "0001",
                "member_name": "DOE, PATIENT",
                "member_id": "000123450",
                "service_date": "20260412",
                "status": "valid",
                "error_count": 0,
                "warning_count": 0,
                "issues": [],
            },
            {
                "index": 1,
                "transaction_control_number": "0002",
                "member_name": "DOE, JANE",
                "member_id": "000123451",
                "service_date": "20260412",
                "status": "invalid",
                "error_count": 1,
                "warning_count": 0,
                "issues": [
                    {
                        "severity": "error",
                        "level": "dc_medicaid",
                        "code": "DCM_INVALID_PAYER_NAME",
                        "message": "Invalid payer name.",
                        "transaction_index": 1,
                        "transaction_control_number": "0002",
                    }
                ],
            },
            {
                "index": 2,
                "transaction_control_number": "0003",
                "member_name": "DOE, JACK",
                "member_id": "000123452",
                "service_date": "20260412",
                "status": "valid",
                "error_count": 0,
                "warning_count": 0,
                "issues": [],
            },
        ],
        "summary": {
            "total_patients": 3,
            "valid_patients": 2,
            "invalid_patients": 1,
        },
    }
