from __future__ import annotations

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def test_export_xlsx_returns_valid_workbook(client: TestClient) -> None:
    payload = {
        "filename": "eligibility.xlsx",
        "payer_name": "DC MEDICAID",
        "summary": {"total": 1, "active": 1, "inactive": 0, "error": 0, "unknown": 0},
        "results": [
            {
                "member_name": "DOE, JOHN",
                "member_id": "12345678",
                "overall_status": "active",
                "eligibility_segments": [{"eligibility_code": "1", "service_type_code": "30"}],
                "benefit_entities": [],
                "aaa_errors": [],
            }
        ],
    }

    response = client.post("/api/v1/export/xlsx", json=payload)

    assert response.status_code == 200
    workbook = load_workbook(BytesIO(response.content))
    assert workbook.sheetnames == ["Summary", "Eligibility Results"]


def test_templates_and_profiles_endpoints(client: TestClient) -> None:
    csv_template = client.get("/api/v1/templates/eligibility_template.csv")
    xlsx_template = client.get("/api/v1/templates/eligibility_template.xlsx")
    profiles = client.get("/api/v1/profiles")
    defaults = client.get("/api/v1/profiles/dc_medicaid/defaults")
    missing = client.get("/api/v1/profiles/missing/defaults")

    assert csv_template.status_code == 200
    assert csv_template.text.startswith("last_name,first_name")
    assert xlsx_template.status_code == 200
    assert xlsx_template.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert profiles.status_code == 200
    assert any(profile["name"] == "dc_medicaid" for profile in profiles.json()["profiles"])
    assert defaults.status_code == 200
    assert defaults.json()["payer_name"] == "DC MEDICAID"
    assert missing.status_code == 404
