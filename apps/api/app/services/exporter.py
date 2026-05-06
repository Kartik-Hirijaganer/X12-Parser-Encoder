"""Excel export services."""

from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from io import BytesIO
from typing import Any, Protocol, TypeAlias, cast

from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.styles import (  # type: ignore[import-untyped]
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from app.core.logging import get_logger
from app.core.metrics import observe_record_count
from app.schemas.common import EligibilityResult, PlanOption
from app.schemas.export import ExportWorkbookRequest
from app.schemas.validate import ValidateResponse, ValidationSummary
from app.services.plans import PlanView, selected_plan_options, split_plan_description

logger = get_logger(__name__)

HEADER_FILL_RGB = "E8EEF7"
HEADER_FONT_RGB = "111827"
BORDER_RGB = "D1D5DB"
ERROR_ROW_FILL_RGB = "FEF2F2"
ERROR_TEXT_RGB = "991B1B"
SUMMARY_LABEL_FILL_RGB = "F9FAFB"

TEXT_NUMBER_FORMAT = "@"

HEADER_FILL = PatternFill(fill_type="solid", fgColor=HEADER_FILL_RGB)
ERROR_ROW_FILL = PatternFill(fill_type="solid", fgColor=ERROR_ROW_FILL_RGB)
SUMMARY_LABEL_FILL = PatternFill(fill_type="solid", fgColor=SUMMARY_LABEL_FILL_RGB)
THIN_BOTTOM_BORDER = Border(bottom=Side(style="thin", color=BORDER_RGB))

ExcelCell: TypeAlias = Any
ExcelValue: TypeAlias = str | int | bool | None


class WorkbookLike(Protocol):
    active: object | None

    def create_sheet(self, title: str) -> object: ...

    def save(self, filename: BytesIO) -> None: ...


class WorksheetLike(Protocol):
    title: str
    freeze_panes: str | None
    max_row: int
    max_column: int
    column_dimensions: Any
    auto_filter: Any

    def append(self, iterable: Sequence[ExcelValue]) -> None: ...

    def cell(self, *, row: int, column: int) -> ExcelCell: ...

    def iter_rows(
        self,
        *,
        min_row: int,
        max_row: int,
        max_col: int,
    ) -> Iterable[tuple[ExcelCell, ...]]: ...

    def __getitem__(self, key: object) -> tuple[ExcelCell, ...]: ...


@dataclass(frozen=True)
class SheetColumn:
    header: str
    width: int
    wrap: bool = False
    key: str | None = None
    text_format: bool = False


ELIGIBILITY_RESULT_COLUMNS = [
    SheetColumn("Member Name", 24, key="member_name"),
    SheetColumn("Member ID", 16, key="member_id", text_format=True),
    SheetColumn("Status", 14, key="overall_status"),
    SheetColumn("Status Reason", 34, wrap=True, key="status_reason"),
    SheetColumn("Program", 32, wrap=True, key="program_name"),
    SheetColumn("Payer Code", 14, key="payer_code", text_format=True),
    SheetColumn("Coverage Category", 18, key="category"),
    SheetColumn("Billing Note", 30, wrap=True, key="billing_note"),
    SheetColumn("EB01 Codes", 16, wrap=True, key="all_eb01_codes", text_format=True),
    SheetColumn(
        "Service Types",
        20,
        wrap=True,
        key="all_eb03_service_types",
        text_format=True,
    ),
    SheetColumn("Benefit Entities", 28, wrap=True, key="benefit_entity_names"),
    SheetColumn("Contacts", 34, wrap=True, key="contact_summaries"),
    SheetColumn("AAA Codes", 14, key="aaa_codes", text_format=True),
    SheetColumn("ST Control #", 16, key="st_control_number", text_format=True),
    SheetColumn("Trace #", 22, key="trace_number", text_format=True),
]

ALL_PLAN_RESULT_COLUMNS = [
    SheetColumn("Member Name", 24, key="member_name"),
    SheetColumn("Member ID", 16, key="member_id", text_format=True),
    SheetColumn("Status", 14, key="overall_status"),
    SheetColumn("Status Reason", 34, wrap=True, key="status_reason"),
    SheetColumn("Plan View", 22, key="plan_view"),
    SheetColumn("Program", 32, wrap=True, key="program_name"),
    SheetColumn("Payer Code", 14, key="payer_code", text_format=True),
    SheetColumn("Coverage Category", 18, key="category"),
    SheetColumn("EB04", 12, key="insurance_type_code", text_format=True),
    SheetColumn("EB01", 12, key="eligibility_code", text_format=True),
    SheetColumn("Billing Note", 30, wrap=True, key="billing_note"),
    SheetColumn("Service Types", 20, wrap=True, key="all_eb03_service_types", text_format=True),
    SheetColumn("Benefit Entities", 28, wrap=True, key="benefit_entity_names"),
    SheetColumn("Contacts", 34, wrap=True, key="contact_summaries"),
    SheetColumn("AAA Codes", 14, key="aaa_codes", text_format=True),
    SheetColumn("ST Control #", 16, key="st_control_number", text_format=True),
    SheetColumn("Trace #", 22, key="trace_number", text_format=True),
]

ERROR_COLUMNS = [
    SheetColumn("Member Name", 24, key="member_name"),
    SheetColumn("Member ID", 16, key="member_id", text_format=True),
    SheetColumn("Error Type", 14, key="error_type"),
    SheetColumn("AAA Code", 14, key="aaa_code", text_format=True),
    SheetColumn("Error Summary", 34, wrap=True, key="error_summary"),
    SheetColumn("Recommended Action", 38, wrap=True, key="recommended_action"),
    SheetColumn("Follow-up Action", 18, key="follow_up_action_code"),
    SheetColumn("ST Control #", 16, key="st_control_number", text_format=True),
    SheetColumn("Trace #", 22, key="trace_number", text_format=True),
]

PARSER_ISSUE_COLUMNS = [
    SheetColumn("Transaction #", 16, key="transaction_index"),
    SheetColumn(
        "Transaction Control #",
        24,
        key="transaction_control_number",
        text_format=True,
    ),
    SheetColumn("Segment", 12, key="segment_id"),
    SheetColumn("Location", 28, key="location"),
    SheetColumn("Message", 48, wrap=True, key="message"),
    SheetColumn("Severity", 12, key="severity"),
]


def _new_workbook() -> WorkbookLike:
    return cast(WorkbookLike, Workbook())


def _active_worksheet(workbook: WorkbookLike) -> WorksheetLike:
    sheet = workbook.active
    if sheet is None:
        raise RuntimeError("openpyxl workbook has no active worksheet")
    return cast(WorksheetLike, sheet)


def _created_worksheet(workbook: WorkbookLike, title: str) -> WorksheetLike:
    return cast(WorksheetLike, workbook.create_sheet(title))


def build_workbook_bytes(
    payload: ExportWorkbookRequest,
    *,
    correlation_id: str | None = None,
    metrics_path: str = "/api/v1/export/xlsx",
) -> bytes:
    """Render parsed eligibility results into an Excel workbook."""

    workbook = _new_workbook()
    summary_sheet = _active_worksheet(workbook)
    summary_sheet.title = "Summary"
    summary_rows: list[tuple[str, ExcelValue]] = [
        ("Payer", payload.payer_name or ""),
        ("Total", payload.summary.total),
        ("Active", payload.summary.active),
        ("Inactive", payload.summary.inactive),
        ("Error", payload.summary.error),
        ("Not Found", payload.summary.not_found),
        ("Unknown", payload.summary.unknown),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    _style_summary_sheet(summary_sheet, row_count=len(summary_rows))

    error_results = [result for result in payload.results if _is_error_row(result)]
    if error_results:
        errors_sheet = _created_worksheet(workbook, "Errors")
        errors_sheet.append([column.header for column in ERROR_COLUMNS])
        for result in error_results:
            if result.aaa_errors:
                for aaa_error in result.aaa_errors:
                    errors_sheet.append(
                        [
                            result.member_name,
                            result.member_id,
                            "AAA",
                            aaa_error.code,
                            aaa_error.message,
                            aaa_error.suggestion or _fallback_recommended_action(result),
                            aaa_error.follow_up_action_code,
                            result.st_control_number,
                            result.trace_number,
                        ]
                    )
            else:
                errors_sheet.append(
                    [
                        result.member_name,
                        result.member_id,
                        "STATUS",
                        "",
                        result.status_reason or "",
                        _fallback_recommended_action(result),
                        "",
                        result.st_control_number,
                        result.trace_number,
                    ]
                )
        _style_table_header(errors_sheet, font_rgb=ERROR_TEXT_RGB)
        _apply_table_layout(errors_sheet, ERROR_COLUMNS)

    results_sheet = _created_worksheet(workbook, "Eligibility Results")
    result_columns = (
        ALL_PLAN_RESULT_COLUMNS if payload.plan_view == "all" else ELIGIBILITY_RESULT_COLUMNS
    )
    results_sheet.append([column.header for column in result_columns])

    for result in payload.results:
        if payload.plan_view == "all":
            all_options: list[PlanOption | None] = list(selected_plan_options(result, "all"))
            if not all_options:
                all_options = [None]
            for plan_option in all_options:
                _append_all_plan_result_row(results_sheet, result, plan_option)
                if _is_error_row(result):
                    _fill_row(results_sheet[results_sheet.max_row])
        else:
            selected_option = _selected_plan_option(result, payload.plan_view)
            _append_result_row(results_sheet, result, selected_option)
            if _is_error_row(result):
                _fill_row(results_sheet[results_sheet.max_row])
    _style_table_header(results_sheet)
    _apply_table_layout(results_sheet, result_columns)

    issue_count = payload.parser_issue_count or len(payload.parser_issues)
    if issue_count > 0:
        issues_sheet = _created_worksheet(workbook, "Parser Issues")
        issues_sheet.append([column.header for column in PARSER_ISSUE_COLUMNS])
        for issue in payload.parser_issues:
            issues_sheet.append(
                [
                    issue.transaction_index,
                    issue.transaction_control_number,
                    issue.segment_id,
                    issue.location,
                    issue.message,
                    issue.severity,
                ]
            )
        _style_table_header(issues_sheet)
        _apply_table_layout(issues_sheet, PARSER_ISSUE_COLUMNS)

    output = BytesIO()
    workbook.save(output)
    observe_record_count(path=metrics_path, operation="export_rows", count=len(payload.results))
    workbook_bytes = output.getvalue()
    logger.info(
        "export_workbook_completed",
        extra={
            "correlation_id": correlation_id,
            "path": metrics_path,
            "result_count": len(payload.results),
            "workbook_bytes": len(workbook_bytes),
        },
    )
    return workbook_bytes


def build_validation_workbook_bytes(
    payload: ValidateResponse,
    *,
    correlation_id: str | None = None,
    metrics_path: str = "/api/v1/export/validation/xlsx",
) -> bytes:
    """Render validation results into a three-sheet Excel workbook."""

    workbook = _new_workbook()
    summary_sheet = _active_worksheet(workbook)
    summary_sheet.title = "Summary"

    summary = payload.summary or _validation_summary_from_patients(payload)
    summary_rows: list[tuple[str, ExcelValue]] = [
        ("filename", payload.filename),
        ("is_valid", payload.is_valid),
        ("error_count", payload.error_count),
        ("warning_count", payload.warning_count),
        ("total_patients", summary.total_patients),
        ("valid_patients", summary.valid_patients),
        ("invalid_patients", summary.invalid_patients),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    for cell in summary_sheet["A"][: len(summary_rows)]:
        cell.font = Font(bold=True)

    patient_sheet = _created_worksheet(workbook, "Per-Patient")
    patient_headers = [
        "index",
        "transaction_control_number",
        "member_name",
        "member_id",
        "service_date",
        "status",
        "error_count",
        "warning_count",
    ]
    patient_sheet.append(patient_headers)
    for cell in patient_sheet[1]:
        cell.font = Font(bold=True)
    patient_sheet.freeze_panes = "A2"
    for patient in payload.patients:
        patient_sheet.append(
            [
                patient.index,
                patient.transaction_control_number,
                patient.member_name,
                patient.member_id,
                patient.service_date,
                patient.status,
                patient.error_count,
                patient.warning_count,
            ]
        )

    issues_sheet = _created_worksheet(workbook, "Issues")
    issue_headers = [
        "severity",
        "level",
        "code",
        "message",
        "location",
        "segment_id",
        "element",
        "suggestion",
        "profile",
        "transaction_index",
        "transaction_control_number",
    ]
    issues_sheet.append(issue_headers)
    for cell in issues_sheet[1]:
        cell.font = Font(bold=True)
    issues_sheet.freeze_panes = "A2"
    for issue in payload.issues:
        issues_sheet.append(
            [
                issue.severity,
                issue.level,
                issue.code,
                issue.message,
                issue.location,
                issue.segment_id,
                issue.element,
                issue.suggestion,
                issue.profile,
                issue.transaction_index,
                issue.transaction_control_number,
            ]
        )

    output = BytesIO()
    workbook.save(output)
    observe_record_count(
        path=metrics_path,
        operation="validation_export_rows",
        count=len(payload.patients),
    )
    workbook_bytes = output.getvalue()
    logger.info(
        "validation_export_workbook_completed",
        extra={
            "correlation_id": correlation_id,
            "path": metrics_path,
            "patient_count": len(payload.patients),
            "issue_count": len(payload.issues),
            "workbook_bytes": len(workbook_bytes),
        },
    )
    return workbook_bytes


def _validation_summary_from_patients(payload: ValidateResponse) -> ValidationSummary:
    valid_patients = sum(1 for patient in payload.patients if patient.status == "valid")
    invalid_patients = sum(1 for patient in payload.patients if patient.status == "invalid")
    return ValidationSummary(
        total_patients=len(payload.patients),
        valid_patients=valid_patients,
        invalid_patients=invalid_patients,
    )


def _append_result_row(
    sheet: WorksheetLike,
    result: EligibilityResult,
    option: PlanOption | None,
) -> None:
    sheet.append(
        [
            result.member_name,
            result.member_id,
            _display_status(result.overall_status),
            result.status_reason,
            option.program_name if option else "",
            option.payer_code if option else "",
            option.category if option else "",
            _billing_note(result),
            _all_eb01_codes(result),
            _all_eb03_service_types(result),
            _benefit_entity_names(result),
            _contact_summaries(result),
            _aaa_codes(result),
            result.st_control_number,
            result.trace_number,
        ]
    )


def _append_all_plan_result_row(
    sheet: WorksheetLike,
    result: EligibilityResult,
    option: PlanOption | None,
) -> None:
    sheet.append(
        [
            result.member_name,
            result.member_id,
            _display_status(result.overall_status),
            result.status_reason,
            option.label if option else "",
            option.program_name if option else "",
            option.payer_code if option else "",
            option.category if option else "",
            option.insurance_type_code if option else "",
            option.eligibility_code if option else "",
            _billing_note(result),
            _all_eb03_service_types(result),
            _benefit_entity_names(result),
            _contact_summaries(result),
            _aaa_codes(result),
            result.st_control_number,
            result.trace_number,
        ]
    )


def _selected_plan_option(result: EligibilityResult, plan_view: PlanView) -> PlanOption | None:
    options = selected_plan_options(result, plan_view)
    return options[0] if options else None


def _primary_plan_summary(result: EligibilityResult, plan_view: PlanView = "agency") -> str | None:
    option = _selected_plan_option(result, plan_view)
    if option is None:
        return None
    return " | ".join(
        part for part in (option.program_name, option.payer_code, option.category) if part
    )


def _split_plan_description(description: str | None) -> tuple[str, str, str]:
    return split_plan_description(description)


def _display_status(value: str | None) -> str:
    if value is None:
        return ""
    normalized = value.strip()
    if not normalized:
        return ""
    status_labels = {
        "active": "Active",
        "inactive": "Inactive",
        "error": "Error",
        "not_found": "Not Found",
        "unknown": "Unknown",
    }
    return status_labels.get(normalized.lower(), normalized.replace("_", " ").title())


def _billing_note(result: EligibilityResult) -> str:
    if result.status_reason:
        return result.status_reason
    if result.aaa_errors:
        return result.aaa_errors[0].message
    return ""


def _is_error_row(result: EligibilityResult) -> bool:
    return result.overall_status in {"error", "not_found"} or bool(result.aaa_errors)


def _fallback_recommended_action(result: EligibilityResult) -> str:
    if result.overall_status == "not_found":
        return "Confirm the member ID and demographics before resubmitting."
    return "Review the status reason and source 271 response before resubmitting."


def _style_summary_sheet(sheet: WorksheetLike, *, row_count: int) -> None:
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 16
    for row_index in range(1, row_count + 1):
        label_cell = sheet.cell(row=row_index, column=1)
        value_cell = sheet.cell(row=row_index, column=2)
        label_cell.font = Font(bold=True, color=HEADER_FONT_RGB)
        label_cell.fill = SUMMARY_LABEL_FILL
        label_cell.alignment = Alignment(vertical="top")
        value_cell.alignment = Alignment(vertical="top")


def _style_table_header(sheet: WorksheetLike, *, font_rgb: str = HEADER_FONT_RGB) -> None:
    for cell in sheet[1]:
        cell.font = Font(bold=True, color=font_rgb)
        cell.fill = HEADER_FILL
        cell.border = THIN_BOTTOM_BORDER
        cell.alignment = Alignment(vertical="center")


def _apply_table_layout(sheet: WorksheetLike, column_specs: list[SheetColumn]) -> None:
    for column_index, column in enumerate(column_specs, start=1):
        column_letter = get_column_letter(column_index)
        sheet.column_dimensions[column_letter].width = column.width

    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        last_column = get_column_letter(sheet.max_column)
        sheet.auto_filter.ref = f"A1:{last_column}{sheet.max_row}"

    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, max_col=len(column_specs)):
        for column_index, cell in enumerate(row):
            column = column_specs[column_index]
            cell.alignment = Alignment(vertical="top", wrap_text=column.wrap)
            if column.text_format:
                cell.number_format = TEXT_NUMBER_FORMAT


def _fill_row(row: Iterable[ExcelCell]) -> None:
    for cell in row:
        cell.fill = ERROR_ROW_FILL


def _all_eb01_codes(result: EligibilityResult) -> str:
    return ", ".join(segment.eligibility_code for segment in result.eligibility_segments)


def _all_eb03_service_types(result: EligibilityResult) -> str:
    service_types: list[str] = []
    for segment in result.eligibility_segments:
        if segment.service_type_codes:
            service_types.extend(segment.service_type_codes)
        elif segment.service_type_code:
            service_types.append(segment.service_type_code)
    return ", ".join(service_types)


def _benefit_entity_names(result: EligibilityResult) -> str:
    names: list[str] = []
    for entity in result.benefit_entities:
        name = entity.name or entity.description or entity.identifier
        if name:
            names.append(name)
    return ", ".join(names)


def _contact_summaries(result: EligibilityResult) -> str:
    contacts: list[str] = []
    for entity in result.benefit_entities:
        contacts.extend(entity.contacts)
    return ", ".join(contacts)


def _aaa_codes(result: EligibilityResult) -> str:
    return ", ".join(error.code for error in result.aaa_errors)
