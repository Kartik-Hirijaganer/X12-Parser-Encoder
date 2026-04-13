"""Patient-template normalization helpers."""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO, StringIO
from typing import Any

from fastapi import HTTPException, status
from openpyxl import load_workbook  # type: ignore[import-untyped]
from x12_edi_tools.payers.dc_medicaid.constants import VALID_SERVICE_TYPE_CODES

from app.schemas.common import Correction, PatientRecord, RowError, WarningMessage

CANONICAL_COLUMNS = [
    "last_name",
    "first_name",
    "date_of_birth",
    "gender",
    "member_id",
    "ssn",
    "service_type_code",
    "service_date",
    "service_date_end",
]
REQUIRED_COLUMNS = {
    "last_name",
    "first_name",
    "date_of_birth",
    "gender",
    "service_date",
}
_CAMEL_TO_SNAKE = {
    "dateOfBirth": "date_of_birth",
    "firstName": "first_name",
    "lastName": "last_name",
    "memberId": "member_id",
    "serviceDate": "service_date",
    "serviceDateEnd": "service_date_end",
    "serviceTypeCode": "service_type_code",
}


@dataclass(slots=True)
class NormalizedPatients:
    """Normalized patient collection plus corrections, warnings, and errors."""

    patients: list[PatientRecord] = field(default_factory=list)
    corrections: list[Correction] = field(default_factory=list)
    warnings: list[WarningMessage] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)


def parse_tabular_bytes(content: bytes, extension: str) -> tuple[list[str], list[dict[str, Any]]]:
    """Parse xlsx/csv/tsv/txt content into canonical header/value mappings."""

    if extension == ".xlsx":
        return _parse_xlsx(content)
    return _parse_delimited(content, extension)


def normalize_patient_rows(
    rows: list[dict[str, Any]],
    *,
    default_service_type_code: str,
    default_service_date: str | None,
    row_offset: int,
) -> NormalizedPatients:
    """Normalize raw row dictionaries into patient records."""

    batch = NormalizedPatients()
    for index, row in enumerate(rows, start=row_offset):
        result = normalize_patient_row(
            row,
            row_number=index,
            default_service_type_code=default_service_type_code,
            default_service_date=default_service_date,
        )
        if result.patient is not None:
            batch.patients.append(result.patient)
        batch.corrections.extend(result.corrections)
        batch.warnings.extend(result.warnings)
        batch.errors.extend(result.errors)
    return batch


@dataclass(slots=True)
class NormalizedPatientRow:
    patient: PatientRecord | None
    corrections: list[Correction] = field(default_factory=list)
    warnings: list[WarningMessage] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)


def normalize_patient_row(
    row: dict[str, Any],
    *,
    row_number: int,
    default_service_type_code: str,
    default_service_date: str | None,
) -> NormalizedPatientRow:
    """Normalize one patient record from uploaded/template data or JSON input."""

    result = NormalizedPatientRow(patient=None)
    normalized_row = {_normalize_header(key): value for key, value in row.items()}

    def read_value(field: str) -> str | None:
        value = normalized_row.get(field)
        if value is None:
            return None
        rendered = str(value).strip()
        return rendered or None

    last_name = _normalize_name(
        read_value("last_name"),
        field_name="last_name",
        row_number=row_number,
        corrections=result.corrections,
    )
    first_name = _normalize_name(
        read_value("first_name"),
        field_name="first_name",
        row_number=row_number,
        corrections=result.corrections,
    )
    date_of_birth = _normalize_date_field(
        read_value("date_of_birth"),
        field_name="date_of_birth",
        row_number=row_number,
        corrections=result.corrections,
    )
    gender = _normalize_gender(
        read_value("gender"),
        row_number=row_number,
        corrections=result.corrections,
    )
    member_id = _normalize_identifier(
        read_value("member_id"),
        field_name="member_id",
        row_number=row_number,
        corrections=result.corrections,
        warnings=result.warnings,
    )
    ssn = _normalize_identifier(
        read_value("ssn"),
        field_name="ssn",
        row_number=row_number,
        corrections=result.corrections,
        warnings=None,
    )
    service_type_code = read_value("service_type_code")
    service_date = read_value("service_date")
    service_date_end = read_value("service_date_end")

    if not last_name:
        result.errors.append(
            RowError(row=row_number, field="last_name", message="last_name is required.")
        )
    if not first_name:
        result.errors.append(
            RowError(row=row_number, field="first_name", message="first_name is required.")
        )
    if not date_of_birth:
        result.errors.append(
            RowError(
                row=row_number,
                field="date_of_birth",
                message="date_of_birth is required and must be parseable.",
            )
        )
    if not gender:
        result.errors.append(
            RowError(
                row=row_number,
                field="gender",
                message="gender is required and must be one of M, F, or U.",
            )
        )

    if not member_id and not ssn:
        result.errors.append(
            RowError(
                row=row_number,
                field="member_id",
                message="Provide at least one subscriber identifier: member_id or ssn.",
                suggestion="Add the member ID or SSN column value for this row.",
            )
        )

    if not service_type_code:
        service_type_code = default_service_type_code
        result.corrections.append(
            Correction(
                row=row_number,
                field="service_type_code",
                original_value=None,
                corrected_value=service_type_code,
                message="Filled missing service_type_code from configuration defaults.",
            )
        )
    else:
        corrected_service_type = service_type_code.upper()
        if corrected_service_type != service_type_code:
            result.corrections.append(
                Correction(
                    row=row_number,
                    field="service_type_code",
                    original_value=service_type_code,
                    corrected_value=corrected_service_type,
                    message="Uppercased service_type_code.",
                )
            )
        service_type_code = corrected_service_type

    if service_type_code not in VALID_SERVICE_TYPE_CODES:
        result.errors.append(
            RowError(
                row=row_number,
                field="service_type_code",
                message=f"Unsupported service_type_code '{service_type_code}'.",
                suggestion=(
                    "Use one of the payer-supported service type codes from the template spec."
                ),
            )
        )

    normalized_service_date = _normalize_date_field(
        service_date,
        field_name="service_date",
        row_number=row_number,
        corrections=result.corrections,
    )
    if not normalized_service_date and default_service_date:
        normalized_service_date = default_service_date
        result.corrections.append(
            Correction(
                row=row_number,
                field="service_date",
                original_value=service_date,
                corrected_value=normalized_service_date,
                message="Filled missing service_date from configuration defaults.",
            )
        )
    if not normalized_service_date:
        result.errors.append(
            RowError(
                row=row_number,
                field="service_date",
                message="service_date is required and must be parseable.",
            )
        )

    normalized_service_date_end = _normalize_date_field(
        service_date_end,
        field_name="service_date_end",
        row_number=row_number,
        corrections=result.corrections,
    )

    if result.errors:
        return result

    assert last_name is not None
    assert first_name is not None
    assert date_of_birth is not None
    assert gender is not None
    assert normalized_service_date is not None
    result.patient = PatientRecord(
        last_name=last_name,
        first_name=first_name,
        date_of_birth=date_of_birth,
        gender=gender,
        member_id=member_id,
        ssn=ssn,
        service_type_code=service_type_code,
        service_date=normalized_service_date,
        service_date_end=normalized_service_date_end,
    )
    return result


def validate_template_headers(headers: list[str]) -> list[WarningMessage]:
    """Validate template headers and return non-fatal warnings."""

    normalized_headers = [_normalize_header(header) for header in headers if header]
    missing_headers = sorted(REQUIRED_COLUMNS - set(normalized_headers))
    if missing_headers:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail={
                "message": "Missing required template columns.",
                "missing_columns": missing_headers,
                "required_columns": sorted(REQUIRED_COLUMNS),
            },
        )

    warnings: list[WarningMessage] = []
    extra_headers = sorted(set(normalized_headers) - set(CANONICAL_COLUMNS))
    for column in extra_headers:
        warnings.append(
            WarningMessage(
                field=column,
                message=f"Ignored unrecognized column '{column}'.",
            )
        )
    return warnings


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows = []
    for raw_row in rows[1:]:
        if _row_is_empty(raw_row):
            continue
        data_rows.append(dict(zip(headers, raw_row, strict=False)))
    return list(headers), data_rows


def _parse_delimited(content: bytes, extension: str) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    delimiter = "\t" if extension == ".tsv" else ","
    if extension == ".txt":
        try:
            dialect = csv.Sniffer().sniff(
                text.splitlines()[0] if text.strip() else "",
                delimiters=",\t",
            )
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    rows: list[dict[str, Any]] = []
    for row in reader:
        if _row_is_empty(row.values()):
            continue
        rows.append(dict(row))
    return list(headers), rows


def _row_is_empty(values: Any) -> bool:
    return not any(value not in (None, "") and str(value).strip() for value in values)


def _normalize_header(value: str) -> str:
    snake_case = _CAMEL_TO_SNAKE.get(value, value)
    return snake_case.strip().lower().replace(" ", "_").replace("-", "_")


def _normalize_name(
    value: str | None,
    *,
    field_name: str,
    row_number: int,
    corrections: list[Correction],
) -> str | None:
    if value is None:
        return None
    corrected = value.strip().upper()
    if corrected != value:
        corrections.append(
            Correction(
                row=row_number,
                field=field_name,
                original_value=value,
                corrected_value=corrected,
                message=f"Normalized {field_name} to uppercase.",
            )
        )
    return corrected or None


def _normalize_gender(
    value: str | None,
    *,
    row_number: int,
    corrections: list[Correction],
) -> str | None:
    if value is None:
        return None
    corrected = value.strip().upper()
    if corrected != value:
        corrections.append(
            Correction(
                row=row_number,
                field="gender",
                original_value=value,
                corrected_value=corrected,
                message="Normalized gender casing.",
            )
        )
    if corrected not in {"M", "F", "U"}:
        return None
    return corrected


def _normalize_identifier(
    value: str | None,
    *,
    field_name: str,
    row_number: int,
    corrections: list[Correction],
    warnings: list[WarningMessage] | None,
) -> str | None:
    if value is None:
        return None
    corrected = value.strip()
    if corrected != value:
        corrections.append(
            Correction(
                row=row_number,
                field=field_name,
                original_value=value,
                corrected_value=corrected,
                message=f"Trimmed whitespace from {field_name}.",
            )
        )
    if field_name == "member_id" and warnings is not None and len(corrected) < 8:
        warnings.append(
            WarningMessage(
                row=row_number,
                field=field_name,
                message=(f"Member ID '{corrected}' appears short — DC Medicaid requires 8 digits."),
                suggestion=(
                    f"Confirm whether '{corrected}0' is the intended padded member ID before "
                    "generation."
                ),
            )
        )
    return corrected or None


def _normalize_date_field(
    value: str | None,
    *,
    field_name: str,
    row_number: int,
    corrections: list[Correction],
) -> str | None:
    if value is None:
        return None

    raw = value.strip()
    if not raw:
        return None

    normalized = _parse_date(raw)
    if normalized is None:
        return None
    if normalized != raw:
        corrections.append(
            Correction(
                row=row_number,
                field=field_name,
                original_value=raw,
                corrected_value=normalized,
                message=f"Normalized {field_name} to YYYYMMDD.",
            )
        )
    return normalized


def _parse_date(value: str) -> str | None:
    if len(value) == 8 and value.isdigit():
        try:
            return datetime.strptime(value, "%Y%m%d").strftime("%Y%m%d")
        except ValueError:
            return None

    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%-m/%-d/%Y", "%-m/%-d/%y"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y%m%d")
        except ValueError:
            continue
    return None
