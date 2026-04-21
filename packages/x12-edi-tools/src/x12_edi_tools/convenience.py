"""High-level convenience API for agents and simple integrations.

This module exposes the four functions documented in the public surface:

* :func:`from_csv` / :func:`from_excel` — read a canonical patient template
  (CSV, TSV, or XLSX) into validated :class:`PatientRecord` objects while
  auto-correcting dates, names, and whitespace and flagging non-PHI warnings.
* :func:`build_270` — turn a list of patient records plus a
  :class:`SubmitterConfig` into a typed :class:`Interchange` ready to encode.
* :func:`read_271` — parse a 271 response into dashboard-oriented
  :class:`EligibilityResult` objects (one per subscriber loop) with an
  optional ``to_dataframe()`` projection for analysis.

The implementation deliberately keeps the library independent of the FastAPI
layer: everything below only uses the stdlib plus Pydantic and the library's
own models. Excel and pandas imports are gated behind optional extras.
"""

from __future__ import annotations

import csv
from collections.abc import Iterable, Iterator, Mapping, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from itertools import count
from pathlib import Path
from typing import Any

from pydantic import BaseModel, ConfigDict, Field

from x12_edi_tools.common.delimiters import Delimiters
from x12_edi_tools.common.enums import (
    AcknowledgmentRequested,
    EntityIdentifierCode,
    GenderCode,
    HierarchicalLevelCode,
    ServiceTypeCode,
    UsageIndicator,
)
from x12_edi_tools.common.types import PathLikeStr
from x12_edi_tools.config import SubmitterConfig
from x12_edi_tools.exceptions import (
    ConfigurationError,
    X12EncodeError,
    X12ParseError,
    X12ValidationError,
)
from x12_edi_tools.models import (
    BHTSegment,
    DMGSegment,
    DTPSegment,
    EQSegment,
    FunctionalGroup,
    GESegment,
    GSSegment,
    HLSegment,
    IEASegment,
    Interchange,
    ISASegment,
    Loop2000A_270,
    Loop2000B_270,
    Loop2000C_270,
    Loop2100A_270,
    Loop2100B_270,
    Loop2100C_270,
    Loop2110C_270,
    NM1Segment,
    PERSegment,
    PRVSegment,
    REFSegment,
    SESegment,
    STSegment,
    Transaction270,
    Transaction271,
    TRNSegment,
)
from x12_edi_tools.models.loops import Loop2000B_271, Loop2000C_271, Loop2110C_271
from x12_edi_tools.parser import ParseResult, parse
from x12_edi_tools.payers import get_profile
from x12_edi_tools.payers.dc_medicaid.constants import (
    AAA_REASON_MESSAGES,
    AAA_REASON_SUGGESTIONS,
    VALID_SERVICE_TYPE_CODES,
)
from x12_edi_tools.validator.base import as_list

__all__ = [
    "AAAError",
    "BenefitEntity",
    "Correction",
    "EligibilityResult",
    "EligibilitySegment",
    "ImportResult",
    "PatientRecord",
    "RowError",
    "WarningMessage",
    "build_270",
    "from_csv",
    "from_excel",
    "read_271",
]

# --- Implementation reference ------------------------------------------------

IMPLEMENTATION_REFERENCE = "005010X279A1"

_CANONICAL_COLUMNS: tuple[str, ...] = (
    "last_name",
    "first_name",
    "date_of_birth",
    "gender",
    "member_id",
    "ssn",
    "service_type_code",
    "service_date",
    "service_date_end",
)
_REQUIRED_COLUMNS = frozenset(
    {"last_name", "first_name", "date_of_birth", "gender", "service_date"}
)
_CAMEL_TO_SNAKE = {
    "dateOfBirth": "date_of_birth",
    "firstName": "first_name",
    "lastName": "last_name",
    "memberId": "member_id",
    "serviceDate": "service_date",
    "serviceDateEnd": "service_date_end",
    "serviceTypeCode": "service_type_code",
}

_ACTIVE_CODES = frozenset({"1", "2", "3", "4", "5"})
_INACTIVE_CODES = frozenset({"6", "7", "8"})


# --- Public patient + import result types -----------------------------------


class PatientRecord(BaseModel):
    """A single subscriber inquiry row after auto-correction."""

    model_config = ConfigDict(str_strip_whitespace=True)

    last_name: str = Field(min_length=1)
    first_name: str = Field(min_length=1)
    date_of_birth: str = Field(min_length=8, max_length=8)
    gender: str = Field(pattern="^(M|F|U)$")
    member_id: str | None = None
    ssn: str | None = None
    service_type_code: str = Field(default="30", min_length=1)
    service_date: str = Field(min_length=8, max_length=8)
    service_date_end: str | None = None


@dataclass(slots=True)
class Correction:
    """A silent auto-correction applied during import."""

    row: int
    field: str
    original_value: str | None
    corrected_value: str | None
    message: str


@dataclass(slots=True)
class WarningMessage:
    """A non-fatal warning that requires user confirmation (never silently applied)."""

    row: int | None
    field: str | None
    message: str
    suggestion: str | None = None


@dataclass(slots=True)
class RowError:
    """A per-row validation failure returned alongside valid rows."""

    row: int
    field: str | None
    message: str
    suggestion: str | None = None


@dataclass(slots=True)
class ImportResult:
    """Return value of :func:`from_csv` and :func:`from_excel`.

    The list is iterable so existing ``list[PatientRecord]`` call sites keep
    working without changes; callers that want full partial-result metadata
    (``corrections``, ``warnings``, ``errors``) can access the attributes
    directly.
    """

    patients: list[PatientRecord] = field(default_factory=list)
    corrections: list[Correction] = field(default_factory=list)
    warnings: list[WarningMessage] = field(default_factory=list)
    errors: list[RowError] = field(default_factory=list)

    def __iter__(self) -> Iterator[PatientRecord]:
        return iter(self.patients)

    def __len__(self) -> int:
        return len(self.patients)

    def __getitem__(self, index: int) -> PatientRecord:
        return self.patients[index]

    def __bool__(self) -> bool:
        return bool(self.patients)


# --- Public 271 result types -------------------------------------------------


@dataclass(slots=True)
class EligibilitySegment:
    """One EB segment projected onto dashboard-friendly fields."""

    eligibility_code: str
    service_type_code: str | None = None
    coverage_level_code: str | None = None
    insurance_type_code: str | None = None
    plan_coverage_description: str | None = None
    monetary_amount: str | None = None
    quantity: str | None = None
    in_plan_network_indicator: str | None = None


@dataclass(slots=True)
class BenefitEntity:
    """Supplemental entity/benefit metadata from REF segments inside a 2110C loop."""

    loop_identifier: str | None
    qualifier: str | None
    identifier: str | None
    description: str | None


@dataclass(slots=True)
class AAAError:
    """An AAA reject reason code mapped to a plain-English message."""

    code: str
    message: str
    follow_up_action_code: str | None = None
    suggestion: str | None = None


@dataclass(slots=True)
class EligibilityResult:
    """One subscriber's eligibility projection from a 271 transaction."""

    member_name: str
    member_id: str | None
    overall_status: str
    eligibility_segments: list[EligibilitySegment] = field(default_factory=list)
    benefit_entities: list[BenefitEntity] = field(default_factory=list)
    aaa_errors: list[AAAError] = field(default_factory=list)


@dataclass(slots=True)
class EligibilityResultSet:
    """Structured 271 output returned by :func:`read_271`."""

    payer_name: str | None
    transaction_count: int
    results: list[EligibilityResult] = field(default_factory=list)
    parse_errors: list[Any] = field(default_factory=list)

    # Iteration/length helpers make ``for r in read_271(...)`` feel natural.
    def __iter__(self) -> Iterator[EligibilityResult]:
        return iter(self.results)

    def __len__(self) -> int:
        return len(self.results)

    def __getitem__(self, index: int) -> EligibilityResult:
        return self.results[index]

    @property
    def summary(self) -> dict[str, int]:
        """Counts by overall_status bucket."""

        counts = {"total": len(self.results), "active": 0, "inactive": 0, "error": 0, "unknown": 0}
        for result in self.results:
            counts[result.overall_status] = counts.get(result.overall_status, 0) + 1
        return counts

    def to_dataframe(self) -> object:
        """Project one row per subscriber into a pandas DataFrame.

        Requires the optional ``pandas`` extra. Raises
        :class:`X12ValidationError` when pandas is not installed so callers
        can prompt installation without catching ``ImportError`` directly.
        """

        try:
            import pandas as pd  # type: ignore[import-untyped]
        except ModuleNotFoundError as exc:  # pragma: no cover - exercised via extras
            raise X12ValidationError(
                "EligibilityResultSet.to_dataframe() requires the optional pandas extra. "
                "Install with `pip install x12-edi-tools[pandas]`.",
            ) from exc

        rows: list[dict[str, Any]] = []
        for result in self.results:
            primary_eb = result.eligibility_segments[0] if result.eligibility_segments else None
            rows.append(
                {
                    "member_name": result.member_name,
                    "member_id": result.member_id,
                    "overall_status": result.overall_status,
                    "eligibility_code": primary_eb.eligibility_code if primary_eb else None,
                    "service_type_code": primary_eb.service_type_code if primary_eb else None,
                    "insurance_type_code": primary_eb.insurance_type_code if primary_eb else None,
                    "plan_coverage_description": (
                        primary_eb.plan_coverage_description if primary_eb else None
                    ),
                    "monetary_amount": primary_eb.monetary_amount if primary_eb else None,
                    "benefit_entity_count": len(result.benefit_entities),
                    "aaa_error_codes": ",".join(error.code for error in result.aaa_errors) or None,
                    "eligibility_segment_count": len(result.eligibility_segments),
                }
            )
        return pd.DataFrame(rows)


# --- Import: CSV / Excel / TSV ----------------------------------------------


def from_csv(
    path: PathLikeStr,
    *,
    default_service_type_code: str = "30",
    default_service_date: str | None = None,
) -> ImportResult:
    """Load patient records from a canonical CSV (or TSV) template.

    The template must expose the canonical snake_case columns documented in
    ``apps/api/templates/template_spec.md``. Date values in any of
    ``YYYYMMDD``, ``YYYY-MM-DD``, ``MM/DD/YYYY``, or ``MM/DD/YY`` are
    normalized to ``YYYYMMDD``. Names are uppercased; whitespace is trimmed.
    Short member IDs surface as warnings — never silent corrections.
    """

    resolved_path = _resolve_input_path(path)
    extension = resolved_path.suffix.lower() or ".csv"
    headers, rows = _parse_tabular_bytes(resolved_path.read_bytes(), extension)
    return _normalize_rows(
        headers,
        rows,
        default_service_type_code=default_service_type_code,
        default_service_date=default_service_date,
    )


def from_excel(
    path: PathLikeStr,
    *,
    default_service_type_code: str = "30",
    default_service_date: str | None = None,
) -> ImportResult:
    """Load patient records from a canonical Excel (``.xlsx``) template."""

    resolved_path = _resolve_input_path(path)
    extension = resolved_path.suffix.lower() or ".xlsx"
    if extension != ".xlsx":  # pragma: no cover - defensive guard
        raise X12ValidationError(
            f"from_excel() expected a .xlsx file, got '{extension}'. "
            "Use from_csv() for delimited data."
        )
    headers, rows = _parse_tabular_bytes(resolved_path.read_bytes(), extension)
    return _normalize_rows(
        headers,
        rows,
        default_service_type_code=default_service_type_code,
        default_service_date=default_service_date,
    )


# --- Build: 270 interchange --------------------------------------------------


def build_270(
    patients: Sequence[PatientRecord | Mapping[str, Any]] | ImportResult,
    *,
    config: SubmitterConfig,
    profile: str = "dc_medicaid",
    generated_at: datetime | None = None,
) -> Interchange:
    """Build a typed 270 :class:`Interchange` from validated patient records.

    The caller is responsible for ensuring ``patients`` is non-empty; the
    function raises :class:`X12EncodeError` otherwise (the encoder itself
    rejects empty interchanges, but we fail fast here with a better message).

    When ``len(patients) > config.max_batch_size`` a single interchange still
    contains every patient — callers that need auto-splitting should pass the
    resulting interchange list to :func:`encode` via ``build_270`` iterated
    per-batch, or use the web API's ``/generate`` which applies the
    split-at-max-batch-size policy.
    """

    resolved_patients = _coerce_patients(patients)
    if not resolved_patients:
        raise X12EncodeError("build_270() requires at least one patient record")

    payer_profile = get_profile(profile)
    defaults = payer_profile.get_defaults()
    payer_name = str(defaults.get("payer_name", config.payer_name))
    payer_id = str(defaults.get("payer_id", config.payer_id))

    now = generated_at or datetime.now()
    date_yymmdd = now.strftime("%y%m%d")
    date_yyyymmdd = now.strftime("%Y%m%d")
    time_hhmm = now.strftime("%H%M")

    transaction_numbers = count(start=1)
    trace_numbers = count(start=1)
    transactions = [
        _build_transaction(
            patient=patient,
            transaction_number=next(transaction_numbers),
            trace_number=next(trace_numbers),
            config=config,
            payer_name=payer_name,
            payer_id=payer_id,
            date_yyyymmdd=date_yyyymmdd,
            time_hhmm=time_hhmm,
        )
        for patient in resolved_patients
    ]

    isa_control_number = _format_isa_control(config.isa_control_number_start or 1)
    gs_control_number = str(config.gs_control_number_start or 1)
    isa = ISASegment(
        authorization_information_qualifier="00",
        authorization_information=" " * 10,
        security_information_qualifier="00",
        security_information=" " * 10,
        sender_id_qualifier=config.sender_id_qualifier,
        sender_id=f"{config.trading_partner_id:<15}"[:15],
        receiver_id_qualifier=config.receiver_id_qualifier,
        receiver_id=f"{config.interchange_receiver_id:<15}"[:15],
        interchange_date=date_yymmdd,
        interchange_time=time_hhmm,
        repetition_separator="^",
        control_version_number="00501",
        interchange_control_number=isa_control_number,
        acknowledgment_requested=AcknowledgmentRequested(config.acknowledgment_requested),
        usage_indicator=UsageIndicator(config.usage_indicator),
        component_element_separator=":",
    )
    gs = GSSegment(
        functional_identifier_code="HS",
        application_sender_code=config.trading_partner_id,
        application_receiver_code=config.payer_id,
        date=date_yyyymmdd,
        time=time_hhmm,
        group_control_number=gs_control_number,
        responsible_agency_code="X",
        version_release_industry_identifier_code=IMPLEMENTATION_REFERENCE,
    )
    ge = GESegment(
        number_of_transaction_sets_included=len(transactions),
        group_control_number=gs_control_number,
    )
    iea = IEASegment(
        number_of_included_functional_groups=1,
        interchange_control_number=isa_control_number,
    )
    return Interchange(
        isa=isa,
        functional_groups=[FunctionalGroup(gs=gs, transactions=list(transactions), ge=ge)],
        iea=iea,
        delimiters=Delimiters(),
    )


# --- Read: 271 response ------------------------------------------------------


def read_271(path_or_string: PathLikeStr | str) -> EligibilityResultSet:
    """Parse a 271 response into structured eligibility results.

    ``path_or_string`` can be a filesystem path or the raw X12 payload as a
    string. Parsing runs with ``on_error="collect"`` so one malformed
    transaction does not discard the rest of the batch; collected errors are
    available on ``result.parse_errors``.
    """

    raw = _load_raw_x12(path_or_string)
    try:
        parsed: ParseResult = parse(raw, strict=False, on_error="collect")
    except X12ParseError:
        raise

    results: list[EligibilityResult] = []
    payer_name: str | None = None
    transaction_count = 0

    for group in parsed.interchange.functional_groups:
        for transaction in group.transactions:
            if not isinstance(transaction, Transaction271):
                # The convenience surface is 271-only; raising keeps the
                # contract honest instead of silently returning empty results.
                raise X12ValidationError("read_271() expects a 271 eligibility response payload")
            transaction_count += 1
            if payer_name is None and transaction.loop_2000a.loop_2100a is not None:
                payer_name = transaction.loop_2000a.loop_2100a.nm1.last_name
            for receiver_loop in as_list(transaction.loop_2000a.loop_2000b):
                if isinstance(receiver_loop, Loop2000B_271):
                    results.extend(_project_receiver_loop(receiver_loop))

    return EligibilityResultSet(
        payer_name=payer_name,
        transaction_count=transaction_count,
        results=results,
        parse_errors=list(parsed.errors),
    )


# --- Internal: import helpers ------------------------------------------------


def _resolve_input_path(path: PathLikeStr) -> Path:
    resolved = Path(path)
    if not resolved.exists():
        raise X12ValidationError(f"Import source does not exist: {resolved}")
    if not resolved.is_file():
        raise X12ValidationError(f"Import source is not a file: {resolved}")
    return resolved


def _parse_tabular_bytes(
    content: bytes,
    extension: str,
) -> tuple[list[str], list[dict[str, Any]]]:
    """Return ``(headers, rows)`` for supported tabular file formats."""

    if extension == ".xlsx":
        return _parse_xlsx(content)
    return _parse_delimited(content, extension)


def _parse_xlsx(content: bytes) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ModuleNotFoundError as exc:
        raise X12ValidationError(
            "from_excel() requires the optional openpyxl extra. "
            "Install with `pip install x12-edi-tools[excel]`.",
        ) from exc

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    data_rows: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        if _row_is_empty(raw_row):
            continue
        data_rows.append(dict(zip(headers, raw_row, strict=False)))
    return list(headers), data_rows


def _parse_delimited(content: bytes, extension: str) -> tuple[list[str], list[dict[str, Any]]]:
    text = content.decode("utf-8-sig")
    delimiter = "\t" if extension == ".tsv" else ","
    if extension == ".txt":
        try:
            first_line = text.splitlines()[0] if text.strip() else ""
            dialect = csv.Sniffer().sniff(first_line, delimiters=",\t")
            delimiter = dialect.delimiter
        except csv.Error:
            delimiter = ","

    reader = csv.DictReader(StringIO(text), delimiter=delimiter)
    headers = reader.fieldnames or []
    rows: list[dict[str, Any]] = []
    for row in reader:
        if _row_is_empty(row.values()):
            continue
        rows.append(dict(row))
    return list(headers), rows


def _row_is_empty(values: Iterable[Any]) -> bool:
    return not any(value not in (None, "") and str(value).strip() for value in values)


def _normalize_header(value: str) -> str:
    snake = _CAMEL_TO_SNAKE.get(value, value)
    return snake.strip().lower().replace(" ", "_").replace("-", "_")


def _normalize_rows(
    headers: Sequence[str],
    rows: Sequence[Mapping[str, Any]],
    *,
    default_service_type_code: str,
    default_service_date: str | None,
) -> ImportResult:
    """Template-aware normalization shared by ``from_csv`` and ``from_excel``."""

    result = ImportResult()
    normalized_headers = [_normalize_header(header) for header in headers if header]
    missing = sorted(_REQUIRED_COLUMNS - set(normalized_headers))
    if missing:
        raise X12ValidationError(
            "Missing required template columns: "
            + ", ".join(missing)
            + f". Required columns: {', '.join(sorted(_REQUIRED_COLUMNS))}."
        )

    extra = sorted(set(normalized_headers) - set(_CANONICAL_COLUMNS))
    for column in extra:
        result.warnings.append(
            WarningMessage(
                row=None,
                field=column,
                message=f"Ignored unrecognized column '{column}'.",
            )
        )

    for row_index, row in enumerate(rows, start=1):
        _normalize_single_row(
            row,
            row_number=row_index,
            default_service_type_code=default_service_type_code,
            default_service_date=default_service_date,
            result=result,
        )
    return result


def _normalize_single_row(
    row: Mapping[str, Any],
    *,
    row_number: int,
    default_service_type_code: str,
    default_service_date: str | None,
    result: ImportResult,
) -> None:
    normalized = {_normalize_header(str(key)): value for key, value in row.items()}

    def read_field(name: str) -> str | None:
        value = normalized.get(name)
        if value is None:
            return None
        rendered = str(value).strip()
        return rendered or None

    row_errors: list[RowError] = []

    last_name = _normalize_name(
        read_field("last_name"),
        field_name="last_name",
        row_number=row_number,
        corrections=result.corrections,
    )
    first_name = _normalize_name(
        read_field("first_name"),
        field_name="first_name",
        row_number=row_number,
        corrections=result.corrections,
    )
    date_of_birth = _normalize_date(
        read_field("date_of_birth"),
        field_name="date_of_birth",
        row_number=row_number,
        corrections=result.corrections,
    )
    gender = _normalize_gender(
        read_field("gender"),
        row_number=row_number,
        corrections=result.corrections,
    )
    member_id = _normalize_identifier(
        read_field("member_id"),
        field_name="member_id",
        row_number=row_number,
        corrections=result.corrections,
        warnings=result.warnings,
    )
    ssn = _normalize_identifier(
        read_field("ssn"),
        field_name="ssn",
        row_number=row_number,
        corrections=result.corrections,
        warnings=None,
    )

    service_type_code = read_field("service_type_code")
    if service_type_code is None:
        service_type_code = default_service_type_code
        result.corrections.append(
            Correction(
                row=row_number,
                field="service_type_code",
                original_value=None,
                corrected_value=service_type_code,
                message="Filled missing service_type_code from configuration defaults.",
            )
        )
    else:
        upper = service_type_code.upper()
        if upper != service_type_code:
            result.corrections.append(
                Correction(
                    row=row_number,
                    field="service_type_code",
                    original_value=service_type_code,
                    corrected_value=upper,
                    message="Uppercased service_type_code.",
                )
            )
            service_type_code = upper
    if service_type_code not in VALID_SERVICE_TYPE_CODES:
        row_errors.append(
            RowError(
                row=row_number,
                field="service_type_code",
                message=f"Unsupported service_type_code '{service_type_code}'.",
                suggestion=(
                    "Use one of the payer-supported service type codes from the template spec."
                ),
            )
        )

    raw_service_date = read_field("service_date")
    service_date = _normalize_date(
        raw_service_date,
        field_name="service_date",
        row_number=row_number,
        corrections=result.corrections,
    )
    if service_date is None and default_service_date:
        service_date = default_service_date
        result.corrections.append(
            Correction(
                row=row_number,
                field="service_date",
                original_value=raw_service_date,
                corrected_value=service_date,
                message="Filled missing service_date from configuration defaults.",
            )
        )
    service_date_end = _normalize_date(
        read_field("service_date_end"),
        field_name="service_date_end",
        row_number=row_number,
        corrections=result.corrections,
    )

    if last_name is None:
        row_errors.append(RowError(row_number, "last_name", "last_name is required."))
    if first_name is None:
        row_errors.append(RowError(row_number, "first_name", "first_name is required."))
    if date_of_birth is None:
        row_errors.append(
            RowError(
                row_number,
                "date_of_birth",
                "date_of_birth is required and must be parseable.",
            )
        )
    if gender is None:
        row_errors.append(
            RowError(
                row_number,
                "gender",
                "gender is required and must be one of M, F, or U.",
            )
        )
    if service_date is None:
        row_errors.append(
            RowError(
                row_number,
                "service_date",
                "service_date is required and must be parseable.",
            )
        )
    if not member_id and not ssn:
        row_errors.append(
            RowError(
                row_number,
                "member_id",
                "Provide at least one subscriber identifier: member_id or ssn.",
                suggestion="Add the member ID or SSN column value for this row.",
            )
        )

    if row_errors:
        result.errors.extend(row_errors)
        return

    assert last_name is not None
    assert first_name is not None
    assert date_of_birth is not None
    assert gender is not None
    assert service_date is not None
    result.patients.append(
        PatientRecord(
            last_name=last_name,
            first_name=first_name,
            date_of_birth=date_of_birth,
            gender=gender,
            member_id=member_id,
            ssn=ssn,
            service_type_code=service_type_code,
            service_date=service_date,
            service_date_end=service_date_end,
        )
    )


def _normalize_name(
    value: str | None,
    *,
    field_name: str,
    row_number: int,
    corrections: list[Correction],
) -> str | None:
    if value is None:
        return None
    corrected = value.strip().upper()
    if corrected != value:
        corrections.append(
            Correction(
                row=row_number,
                field=field_name,
                original_value=value,
                corrected_value=corrected,
                message=f"Normalized {field_name} to uppercase.",
            )
        )
    return corrected or None


def _normalize_gender(
    value: str | None,
    *,
    row_number: int,
    corrections: list[Correction],
) -> str | None:
    if value is None:
        return None
    corrected = value.strip().upper()
    if corrected != value:
        corrections.append(
            Correction(
                row=row_number,
                field="gender",
                original_value=value,
                corrected_value=corrected,
                message="Normalized gender casing.",
            )
        )
    return corrected if corrected in {"M", "F", "U"} else None


def _normalize_identifier(
    value: str | None,
    *,
    field_name: str,
    row_number: int,
    corrections: list[Correction],
    warnings: list[WarningMessage] | None,
) -> str | None:
    if value is None:
        return None
    corrected = value.strip()
    if corrected != value:
        corrections.append(
            Correction(
                row=row_number,
                field=field_name,
                original_value=value,
                corrected_value=corrected,
                message=f"Trimmed whitespace from {field_name}.",
            )
        )
    if field_name == "member_id" and warnings is not None and 0 < len(corrected) < 8:
        warnings.append(
            WarningMessage(
                row=row_number,
                field=field_name,
                message=(f"Member ID '{corrected}' appears short — DC Medicaid requires 8 digits."),
                suggestion=(
                    f"Confirm whether '{corrected}0' is the intended padded member ID before "
                    "generation."
                ),
            )
        )
    return corrected or None


def _normalize_date(
    value: str | None,
    *,
    field_name: str,
    row_number: int,
    corrections: list[Correction],
) -> str | None:
    if value is None:
        return None
    raw = value.strip()
    if not raw:
        return None
    parsed = _parse_flexible_date(raw)
    if parsed is None:
        return None
    normalized = parsed.strftime("%Y%m%d")
    if normalized != raw:
        corrections.append(
            Correction(
                row=row_number,
                field=field_name,
                original_value=raw,
                corrected_value=normalized,
                message=f"Normalized {field_name} to YYYYMMDD.",
            )
        )
    return normalized


def _parse_flexible_date(value: str) -> date | None:
    if len(value) == 8 and value.isdigit():
        try:
            return datetime.strptime(value, "%Y%m%d").date()
        except ValueError:
            return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


# --- Internal: 270 transaction builder --------------------------------------


def _coerce_patients(
    patients: Sequence[PatientRecord | Mapping[str, Any]] | ImportResult,
) -> list[PatientRecord]:
    if isinstance(patients, ImportResult):
        return list(patients.patients)
    coerced: list[PatientRecord] = []
    for index, item in enumerate(patients):
        if isinstance(item, PatientRecord):
            coerced.append(item)
        elif isinstance(item, Mapping):
            try:
                coerced.append(PatientRecord.model_validate(dict(item)))
            except Exception as exc:
                raise ConfigurationError(
                    f"Patient at index {index} failed validation: {exc}"
                ) from exc
        else:
            raise ConfigurationError(
                f"Patient at index {index} must be a PatientRecord or mapping, "
                f"got {type(item).__name__}"
            )
    return coerced


def _format_isa_control(value: int) -> str:
    return f"{value:09d}"


def _build_transaction(
    *,
    patient: PatientRecord,
    transaction_number: int,
    trace_number: int,
    config: SubmitterConfig,
    payer_name: str,
    payer_id: str,
    date_yyyymmdd: str,
    time_hhmm: str,
) -> Transaction270:
    control_number = f"{transaction_number:04d}"
    subscriber_ref_segments: list[REFSegment] = []
    if patient.ssn:
        subscriber_ref_segments.append(
            REFSegment(
                reference_identification_qualifier="SY",
                reference_identification=patient.ssn,
            )
        )

    dtp_period_format = "D8"
    dtp_period = patient.service_date
    if patient.service_date_end:
        dtp_period_format = "RD8"
        dtp_period = f"{patient.service_date}-{patient.service_date_end}"

    provider_loop = Loop2100B_270(
        nm1=NM1Segment(
            entity_identifier_code=EntityIdentifierCode.PROVIDER,
            entity_type_qualifier=config.provider_entity_type,
            last_name=config.organization_name.upper(),
            id_code_qualifier="XX",
            id_code=config.provider_npi,
        ),
        prv=_provider_prv_segment(config.provider_taxonomy_code),
        per=_provider_per_segment(
            contact_name=config.contact_name,
            contact_phone=config.contact_phone,
            contact_email=config.contact_email,
        ),
    )

    return Transaction270(
        st=STSegment(
            transaction_set_identifier_code="270",
            transaction_set_control_number=control_number,
            implementation_convention_reference=IMPLEMENTATION_REFERENCE,
        ),
        bht=BHTSegment(
            hierarchical_structure_code="0022",
            transaction_set_purpose_code="13",
            reference_identification=f"BATCH{transaction_number:08d}",
            date=date_yyyymmdd,
            time=time_hhmm,
        ),
        loop_2000a=Loop2000A_270(
            hl=HLSegment(
                hierarchical_id_number="1",
                hierarchical_parent_id_number=None,
                hierarchical_level_code=HierarchicalLevelCode.INFORMATION_SOURCE,
                hierarchical_child_code="1",
            ),
            loop_2100a=Loop2100A_270(
                nm1=NM1Segment(
                    entity_identifier_code=EntityIdentifierCode.PAYER,
                    entity_type_qualifier="2",
                    last_name=payer_name,
                    id_code_qualifier="PI",
                    id_code=payer_id,
                )
            ),
            loop_2000b=[
                Loop2000B_270(
                    hl=HLSegment(
                        hierarchical_id_number="2",
                        hierarchical_parent_id_number="1",
                        hierarchical_level_code=HierarchicalLevelCode.INFORMATION_RECEIVER,
                        hierarchical_child_code="1",
                    ),
                    loop_2100b=provider_loop,
                    loop_2000c=[
                        Loop2000C_270(
                            hl=HLSegment(
                                hierarchical_id_number="3",
                                hierarchical_parent_id_number="2",
                                hierarchical_level_code=HierarchicalLevelCode.SUBSCRIBER,
                                hierarchical_child_code="0",
                            ),
                            trn=TRNSegment(
                                trace_type_code="1",
                                reference_identification_1=f"TRACE{trace_number:07d}",
                                originating_company_identifier=config.provider_npi,
                            ),
                            loop_2100c=Loop2100C_270(
                                nm1=NM1Segment(
                                    entity_identifier_code=EntityIdentifierCode.SUBSCRIBER,
                                    entity_type_qualifier="1",
                                    last_name=patient.last_name,
                                    first_name=patient.first_name,
                                    id_code_qualifier="MI" if patient.member_id else None,
                                    id_code=patient.member_id,
                                ),
                                dmg=DMGSegment(
                                    date_time_period_format_qualifier="D8",
                                    date_time_period=patient.date_of_birth,
                                    gender_code=GenderCode(patient.gender),
                                ),
                                ref_segments=subscriber_ref_segments,
                                dtp_segments=[
                                    DTPSegment(
                                        date_time_qualifier="291",
                                        date_time_period_format_qualifier=dtp_period_format,
                                        date_time_period=dtp_period,
                                    )
                                ],
                            ),
                            loop_2110c=[
                                Loop2110C_270(
                                    eq_segments=[
                                        EQSegment(
                                            service_type_code=ServiceTypeCode(
                                                patient.service_type_code
                                            )
                                        )
                                    ],
                                )
                            ],
                        )
                    ],
                )
            ],
        ),
        se=SESegment(
            number_of_included_segments=13,
            transaction_set_control_number=control_number,
        ),
    )


def _provider_prv_segment(taxonomy_code: str | None) -> PRVSegment | None:
    if not taxonomy_code:
        return None
    return PRVSegment(
        provider_code="BI",
        reference_identification_qualifier="PXC",
        reference_identification=taxonomy_code,
    )


def _provider_per_segment(
    *,
    contact_name: str | None,
    contact_phone: str | None,
    contact_email: str | None,
) -> PERSegment | None:
    if not any([contact_name, contact_phone, contact_email]):
        return None
    return PERSegment(
        contact_function_code="IC",
        name=contact_name,
        communication_number_qualifier_1="TE" if contact_phone else None,
        communication_number_1=contact_phone,
        communication_number_qualifier_2="EM" if contact_email else None,
        communication_number_2=contact_email,
    )


# --- Internal: 271 projection helpers ---------------------------------------


def _load_raw_x12(path_or_string: PathLikeStr | str) -> str:
    """Return X12 payload text for either a filesystem path or an inline string."""

    if isinstance(path_or_string, str):
        stripped = path_or_string.lstrip()
        if stripped.startswith("ISA"):
            return path_or_string
        candidate = Path(path_or_string)
        if candidate.exists() and candidate.is_file():
            return candidate.read_text(encoding="utf-8")
        # Fall through: treat as a partial payload; parse() will raise.
        return path_or_string

    candidate = Path(path_or_string)
    if not candidate.exists():
        raise X12ValidationError(f"X12 source does not exist: {candidate}")
    return candidate.read_text(encoding="utf-8")


def _project_receiver_loop(receiver_loop: Loop2000B_271) -> list[EligibilityResult]:
    projections: list[EligibilityResult] = []
    for subscriber_loop in as_list(receiver_loop.loop_2000c):
        if isinstance(subscriber_loop, Loop2000C_271):
            projections.append(_project_subscriber_loop(subscriber_loop))
    return projections


def _project_subscriber_loop(subscriber_loop: Loop2000C_271) -> EligibilityResult:
    eligibility_segments: list[EligibilitySegment] = []
    benefit_entities: list[BenefitEntity] = []
    aaa_errors: list[AAAError] = []
    aaa_errors.extend(_map_aaa_segments(as_list(subscriber_loop.aaa_segments)))
    aaa_errors.extend(
        _map_aaa_segments(as_list(getattr(subscriber_loop.loop_2100c, "aaa_segments", [])))
    )

    for inquiry_loop in as_list(subscriber_loop.loop_2110c):
        if not isinstance(inquiry_loop, Loop2110C_271):
            continue
        eligibility_segments.extend(_map_eligibility_segments(inquiry_loop))
        benefit_entities.extend(_map_benefit_entities(inquiry_loop))
        aaa_errors.extend(_map_aaa_segments(as_list(inquiry_loop.aaa_segments)))

    nm1 = subscriber_loop.loop_2100c.nm1
    member_name = ", ".join(part for part in [nm1.last_name, nm1.first_name] if part)
    return EligibilityResult(
        member_name=member_name or "UNKNOWN",
        member_id=nm1.id_code,
        overall_status=_overall_status(eligibility_segments, aaa_errors),
        eligibility_segments=eligibility_segments,
        benefit_entities=benefit_entities,
        aaa_errors=aaa_errors,
    )


def _map_eligibility_segments(loop: Loop2110C_271) -> list[EligibilitySegment]:
    projected: list[EligibilitySegment] = []
    for eb in loop.eb_segments:
        projected.append(
            EligibilitySegment(
                eligibility_code=str(
                    getattr(
                        eb.eligibility_or_benefit_information,
                        "value",
                        eb.eligibility_or_benefit_information,
                    )
                ),
                service_type_code=_enum_value(eb.service_type_code),
                coverage_level_code=eb.coverage_level_code,
                insurance_type_code=eb.insurance_type_code,
                plan_coverage_description=eb.plan_coverage_description,
                monetary_amount=_decimal_string(eb.monetary_amount),
                quantity=_decimal_string(eb.quantity),
                in_plan_network_indicator=eb.in_plan_network_indicator,
            )
        )
    return projected


def _map_benefit_entities(loop: Loop2110C_271) -> list[BenefitEntity]:
    projected: list[BenefitEntity] = []
    loop_identifier = getattr(loop.ls_segment, "loop_identifier_code", None)
    for ref in loop.ref_segments:
        projected.append(
            BenefitEntity(
                loop_identifier=loop_identifier,
                qualifier=ref.reference_identification_qualifier,
                identifier=ref.reference_identification,
                description=ref.description,
            )
        )
    return projected


def _map_aaa_segments(segments: list[Any]) -> list[AAAError]:
    mapped: list[AAAError] = []
    for segment in segments:
        code = getattr(segment, "reject_reason_code", None)
        if code is None:
            continue
        code_value = str(code)
        mapped.append(
            AAAError(
                code=code_value,
                message=AAA_REASON_MESSAGES.get(
                    code_value,
                    "Eligibility response returned an AAA error.",
                ),
                follow_up_action_code=getattr(segment, "follow_up_action_code", None),
                suggestion=AAA_REASON_SUGGESTIONS.get(code_value),
            )
        )
    return mapped


def _overall_status(
    eligibility_segments: list[EligibilitySegment],
    aaa_errors: list[AAAError],
) -> str:
    if aaa_errors:
        return "error"
    codes = {segment.eligibility_code for segment in eligibility_segments}
    if codes & _ACTIVE_CODES:
        return "active"
    if codes & _INACTIVE_CODES:
        return "inactive"
    return "unknown"


def _decimal_string(value: Decimal | None) -> str | None:
    if value is None:
        return None
    return format(value, "f")


def _enum_value(value: object) -> str | None:
    if value is None:
        return None
    return str(getattr(value, "value", value))
